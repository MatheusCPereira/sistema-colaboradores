#!/usr/bin/env python3
"""
Script de acesso à base de dados.
Chamado pelo servidor Node via child_process.
Recebe JSON no argv[1] como {action, ...params}
Retorna JSON no stdout.
"""

import sys
import json
import os
import hashlib
import uuid
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
EXCEL_FILE = os.path.join(DATA_DIR, 'colaboradores.xlsx')
USERS_FILE = os.path.join(DATA_DIR, 'users.json')

SHEET_NAME = 'BASE'

# Mapeamento interno → coluna Excel
COL_MAP = {
    'seq':           'SEG',
    'nome':          'NOME COMPLETO',
    'matricula':     'MATRICULA',
    'admissao':      'ADMISSÃO',
    'aso':           'DATA DO ASO',
    'pis':           'PIS',
    'cpf':           'CPF',
    'rg':            'RG',
    'expRg':         'Data exp. RG',
    'nascimento':    'Nascimento',
    'naturalidade':  'Naturalidade',
    'nrRegistro':    'Nº REGISTRO',
    'categoria':     'Categoria',
    'vencimentoCnh': 'Vencimento CNH',
    'nrDocumento':   'N° documento',
    'primeiraHab':   '1° Habilitação ',
    'nrSeguranca':   'N° de Segurança',
    'renach':        'RENACH',
    'estadoCnh':     'ESTADO DA CNH',
    'nomePai':       'Nome do Pai',
    'nomeMae':       'Nome da Mãe',
    'funcao':        'FUNÇÃO',
    'demissao':      'DEMISSÃO',
    'endereco':      'ENDEREÇO',
    'numero':        'N.º',
    'bairro':        'BAIRRO',
    'cidade':        'CIDADE',
    'cep':           'CEP',
    'estadoCivil':   'ESTADO CIVIL',
    'telefone':      'TELEFONE',
    'aniversario':   'ANIVERSÁRIANTES',
    'email':         'E-MAIL',
    'horario':       'HORARIO',
    'fixoFolg':      'FIXO/ FOLG',
    'operacao':      'OPERAÇÃO',
    'linkPasta':     'LINK PASTA',
    'status':        'STATUS',
    'genero':        'GENERO',
    'safraNovo':     'SAFRA/ NOVO',
    'uf':            'UF',
}

DATE_KEYS = {'admissao','aso','expRg','nascimento','vencimentoCnh','primeiraHab','demissao'}


def fmt_date(val):
    if val is None or (isinstance(val, float) and __import__('math').isnan(val)):
        return ''
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s in ('nan','NaT','None'):
        return ''
    # tenta parse
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s[:len(fmt)], fmt).strftime('%Y-%m-%d')
        except:
            pass
    return s


def read_excel():
    import pandas as pd
    df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME, dtype=str, keep_default_na=False)
    return df


def rows_to_objects(df):
    import pandas as pd
    result = []
    rev_map = {v: k for k, v in COL_MAP.items()}
    for i, row in enumerate(df.itertuples(index=False)):
        obj = {'_id': i + 1}
        for col in df.columns:
            key = rev_map.get(col)
            if key is None:
                continue
            val = getattr(row, col.replace(' ','_').replace('/','_').replace('.','_').replace('°','').replace('ê','e').replace('ã','a').replace('Ã','A').replace('Ç','C').replace('ç','c').replace('É','E').replace('é','e').replace('Ú','U').replace('ú','u')) if False else row[df.columns.get_loc(col)]
            if key in DATE_KEYS:
                val = fmt_date(val)
            else:
                val = '' if (val is None or str(val) in ('nan','NaT','None','<NA>')) else str(val).strip()
            obj[key] = val
        if obj.get('nome', '').strip():
            result.append(obj)
    return result


def write_excel(objects):
    import pandas as pd
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    # Pegar cabeçalhos existentes
    headers = [cell.value for cell in ws[1]]

    # Limpar dados existentes (manter cabeçalho)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Escrever novos dados
    col_idx = {h: i+1 for i, h in enumerate(headers) if h}

    for row_i, obj in enumerate(objects, start=2):
        for key, col_name in COL_MAP.items():
            if col_name in col_idx:
                ws.cell(row=row_i, column=col_idx[col_name], value=obj.get(key, '') or '')

    wb.save(EXCEL_FILE)


# ===== COLABORADORES =====

def get_colaboradores(search='', status='', page=1, limit=20):
    df = read_excel()
    objects = rows_to_objects(df)

    if search:
        s = search.lower()
        objects = [o for o in objects if
            s in o.get('nome','').lower() or
            s in o.get('matricula','').lower() or
            s in o.get('cpf','').lower() or
            s in o.get('funcao','').lower() or
            s in o.get('cidade','').lower() or
            s in o.get('email','').lower()
        ]
    if status:
        objects = [o for o in objects if o.get('status','').lower() == status.lower()]

    total = len(objects)
    total_pages = max(1, -(-total // limit))  # ceil division
    items = objects[(page-1)*limit : page*limit]
    return {'items': items, 'total': total, 'totalPages': total_pages, 'page': page}


def get_colaborador_by_id(id_):
    df = read_excel()
    objects = rows_to_objects(df)
    for o in objects:
        if o['_id'] == int(id_):
            return o
    return None


def create_colaborador(data):
    df = read_excel()
    objects = rows_to_objects(df)
    new_seq = max([int(o.get('seq') or 0) for o in objects], default=0) + 1
    new_obj = {'_id': len(objects)+1, 'seq': str(new_seq)}
    for k in COL_MAP:
        new_obj[k] = data.get(k, '') or ''
    new_obj['seq'] = str(new_seq)
    objects.append(new_obj)
    write_excel(objects)
    return new_obj


def update_colaborador(id_, data):
    df = read_excel()
    objects = rows_to_objects(df)
    idx = next((i for i,o in enumerate(objects) if o['_id'] == int(id_)), None)
    if idx is None:
        return None
    for k in COL_MAP:
        if k in data:
            objects[idx][k] = data[k] or ''
    write_excel(objects)
    return objects[idx]


def delete_colaborador(id_):
    df = read_excel()
    objects = rows_to_objects(df)
    new_objects = [o for o in objects if o['_id'] != int(id_)]
    if len(new_objects) == len(objects):
        return False
    write_excel(new_objects)
    return True


def get_estatisticas():
    df = read_excel()
    objects = rows_to_objects(df)
    total = len(objects)
    ativos = sum(1 for o in objects if o.get('status','').lower() == 'contratado')
    demitidos = sum(1 for o in objects if o.get('status','').lower() == 'demitido')
    funcoes = {}
    cidades = {}
    generos = {}
    for o in objects:
        f = o.get('funcao','').strip()
        if f: funcoes[f] = funcoes.get(f,0)+1
        c = o.get('cidade','').strip()
        if c: cidades[c] = cidades.get(c,0)+1
        g = o.get('genero','').strip()
        if g: generos[g] = generos.get(g,0)+1
    return {'total':total,'ativos':ativos,'demitidos':demitidos,'funcoes':funcoes,'cidades':cidades,'generos':generos}


# ===== USUÁRIOS =====

def hash_password(pw):
    salt = os.urandom(16).hex()
    h = hashlib.sha256((pw + salt).encode()).hexdigest()
    return f"{salt}${h}"

def verify_password(pw, stored):
    parts = stored.split('$')
    if len(parts) != 2: return False
    salt, h = parts
    return hashlib.sha256((pw + salt).encode()).hexdigest() == h


def read_users():
    if not os.path.exists(USERS_FILE):
        default = [{
            'id': str(uuid.uuid4()),
            'username': 'admin',
            'password': hash_password('admin123'),
            'name': 'Administrador',
            'role': 'admin',
            'createdAt': datetime.now().isoformat()
        }]
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(USERS_FILE, 'w') as f:
            json.dump(default, f, indent=2, ensure_ascii=False)
        return default
    with open(USERS_FILE) as f:
        return json.load(f)

def write_users(users):
    with open(USERS_FILE, 'w') as f:
        json.dump(users, f, indent=2, ensure_ascii=False)


def login(username, password):
    users = read_users()
    for u in users:
        if u['username'] == username:
            if verify_password(password, u['password']):
                safe = {k:v for k,v in u.items() if k != 'password'}
                return {'ok': True, 'user': safe}
    return {'ok': False, 'error': 'Usuário ou senha inválidos'}


def get_all_users():
    return [{k:v for k,v in u.items() if k!='password'} for u in read_users()]


def create_user(data):
    users = read_users()
    if any(u['username'] == data['username'] for u in users):
        return {'error': 'Usuário já existe'}
    new_user = {
        'id': str(uuid.uuid4()),
        'username': data['username'],
        'password': hash_password(data['password']),
        'name': data.get('name',''),
        'role': data.get('role','viewer'),
        'createdAt': datetime.now().isoformat()
    }
    users.append(new_user)
    write_users(users)
    return {k:v for k,v in new_user.items() if k!='password'}


def update_user(id_, data):
    users = read_users()
    idx = next((i for i,u in enumerate(users) if u['id'] == id_), None)
    if idx is None:
        return {'error': 'Usuário não encontrado'}
    for k in ('name','username','role'):
        if k in data:
            users[idx][k] = data[k]
    if data.get('password'):
        users[idx]['password'] = hash_password(data['password'])
    write_users(users)
    return {k:v for k,v in users[idx].items() if k!='password'}


def delete_user(id_):
    users = read_users()
    new_users = [u for u in users if u['id'] != id_]
    if len(new_users) == len(users):
        return False
    write_users(new_users)
    return True


# ===== DISPATCHER =====

def main():
    # Lê do stdin — compatível com Windows, Linux e Mac
    raw = sys.stdin.read().strip()
    if not raw:
        print(json.dumps({'error': 'Nenhum dado recebido no stdin'}))
        return
    req = json.loads(raw)
    action = req['action']
    result = {}

    try:
        if action == 'get_colaboradores':
            result = get_colaboradores(
                search=req.get('search',''),
                status=req.get('status',''),
                page=int(req.get('page',1)),
                limit=int(req.get('limit',20))
            )
        elif action == 'get_colaborador':
            c = get_colaborador_by_id(req['id'])
            result = c if c else {'error': 'Não encontrado'}
        elif action == 'create_colaborador':
            result = create_colaborador(req['data'])
        elif action == 'update_colaborador':
            r = update_colaborador(req['id'], req['data'])
            result = r if r else {'error': 'Não encontrado'}
        elif action == 'delete_colaborador':
            result = {'ok': delete_colaborador(req['id'])}
        elif action == 'get_estatisticas':
            result = get_estatisticas()
        elif action == 'login':
            result = login(req['username'], req['password'])
        elif action == 'get_users':
            result = get_all_users()
        elif action == 'create_user':
            result = create_user(req['data'])
        elif action == 'update_user':
            result = update_user(req['id'], req['data'])
        elif action == 'delete_user':
            result = {'ok': delete_user(req['id'])}
        else:
            result = {'error': f'Ação desconhecida: {action}'}
    except Exception as e:
        import traceback
        result = {'error': str(e), 'trace': traceback.format_exc()}

    print(json.dumps(result, ensure_ascii=False, default=str))


if __name__ == '__main__':
    main()